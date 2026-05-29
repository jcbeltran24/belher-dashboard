#!/usr/bin/env python3
"""
Encrypta el reporte mensual de liquidación Calavo (xlsx → json → AES-GCM)
y lo embebe en api/settle.js como blob base64.

Uso:
  python3 scripts/encrypt-settle.py <ruta.xlsx> [--password XXX]

Si no pasas --password, lee .settle-password (raíz del repo). Si tampoco
existe, genera una contraseña aleatoria de 24 chars y la guarda en
.settle-password (gitignored).

Parámetros criptográficos (deben coincidir con el cliente):
  - KDF: PBKDF2-HMAC-SHA256, 600,000 iteraciones
  - Salt: 16 bytes random por encriptación
  - Cipher: AES-256-GCM
  - IV: 12 bytes random por encriptación
  - Output JSON: { saltB64, ivB64, ctB64, version: 1 }

Convención de hoja:
  - Overview Payments → pagos semanales
  - Beltran Summary Statement → avances y resumen
  - Loan Amortization Schedule → detalle por embarque (Calavo PO)
"""
import argparse, base64, json, os, secrets, string, sys
from pathlib import Path
from datetime import datetime, date

import openpyxl
from openpyxl.utils.datetime import from_excel
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

ROOT = Path(__file__).resolve().parent.parent
PWFILE = ROOT / ".settle-password"
OUT = ROOT / "api" / "settle.js"

PBKDF2_ITERS = 600_000


def jsonable(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10] if isinstance(v, date) and not isinstance(v, datetime) else v.isoformat()
    return v


def xl_date(v):
    """Para columnas-fecha: convierte Excel serial floats a ISO string.
       Strings/None se devuelven tal cual; datetimes pasan por jsonable."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return jsonable(v)
    if isinstance(v, (int, float)):
        # Rango defensivo: serials Excel >= 1 (1900-01-01) y < 80000 (~2119)
        if 1 <= float(v) < 80000:
            try:
                return from_excel(float(v)).isoformat()
            except Exception:
                return str(v)
        return str(v)
    return str(v)


def read_overview_payments(ws):
    """Cols: Week, Received Week, Wire Date, Payment Amount, Units Received, Containers."""
    out = []
    for r in range(8, ws.max_row + 1):
        week = ws.cell(r, 1).value
        if not week or not str(week).strip().lower().startswith("week"):
            continue
        out.append({
            "week": str(week).strip(),
            "received_week": xl_date(ws.cell(r, 2).value),
            "wire_date":     xl_date(ws.cell(r, 3).value),
            "payment_amount": ws.cell(r, 4).value,
            "units_received": ws.cell(r, 5).value,
            "containers":     ws.cell(r, 6).value,
        })
    return out


def read_beltran_summary(ws):
    """Estructura key-value libre; capturamos pares (descripción → monto) y los grupos."""
    out = []
    current_group = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        d = ws.cell(r, 4).value
        if a and (d is None) and isinstance(a, str) and a.strip() and a.strip() not in {"Agricola Belher", "Description"}:
            current_group = a.strip()
            continue
        if a and d is not None:
            desc = str(a).strip()
            if desc.lower() == "description":
                continue
            out.append({
                "group": current_group,
                "description": desc,
                "amount": d,
            })
    return out


def read_loan_amortization(ws):
    """Header en R1: Calavo PO, Vendor, Date Rcvd, Grower Reference, House, Units Recvd,
       Pckgs Invcd, Pckgs Pndng, Rework, Invtry OnHand, Total Revenue, Calavo Comm,
       Pick & Pack, Z-Code Expense, ..."""
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        row = {}
        po = ws.cell(r, 1).value
        if not po or str(po).strip() in {"", "0"}:
            continue
        for i, h in enumerate(headers):
            if not h:
                continue
            val = ws.cell(r, i + 1).value
            # Columnas-fecha: coercionar serials Excel a ISO strings
            if "date" in h.lower() or h.lower().endswith(" rcvd"):
                val = xl_date(val)
            else:
                val = jsonable(val)
            row[h] = val
        out.append(row)
    return out


def extract(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = wb.sheetnames

    def pick(name):
        for s in sheets:
            if s.lower().startswith(name.lower()):
                return wb[s]
        return None

    overview = pick("Overview Payments")
    summary = pick("Beltran Summary Statement")
    loan = pick("Loan Amortization Schedule")

    data = {
        "source_file": xlsx_path.name,
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "overview_payments": read_overview_payments(overview) if overview else [],
        "beltran_summary": read_beltran_summary(summary) if summary else [],
        "embarques": read_loan_amortization(loan) if loan else [],
    }
    return data


def encrypt(data: dict, password: str) -> dict:
    plaintext = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
    salt = secrets.token_bytes(16)
    iv = secrets.token_bytes(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=PBKDF2_ITERS)
    key = kdf.derive(password.encode("utf-8"))
    ct = AESGCM(key).encrypt(iv, plaintext, None)
    return {
        "version": 1,
        "kdf": "PBKDF2-HMAC-SHA256",
        "iters": PBKDF2_ITERS,
        "saltB64": base64.b64encode(salt).decode(),
        "ivB64": base64.b64encode(iv).decode(),
        "ctB64": base64.b64encode(ct).decode(),
    }


def gen_password(n=24) -> str:
    alphabet = string.ascii_letters + string.digits + "-_"
    return "".join(secrets.choice(alphabet) for _ in range(n))


def write_function(blob: dict):
    blob_js = json.dumps(blob, indent=2)
    src = f"""// Auto-generado por scripts/encrypt-settle.py — NO EDITAR A MANO.
// Liquidación Calavo encriptada (AES-256-GCM, PBKDF2-HMAC-SHA256, {PBKDF2_ITERS} iters).
// El password vive en Vercel como env var SETTLE_PASSWORD y NO se loguea.

const BLOB = {blob_js};

const CORS = {{
  'Access-Control-Allow-Origin':  '*',
  'Access-Control-Allow-Methods': 'POST, OPTIONS',
  'Access-Control-Allow-Headers': 'content-type, x-settle-password',
}};

function timingSafeEqualStr(a, b) {{
  if (typeof a !== 'string' || typeof b !== 'string') return false;
  if (a.length !== b.length) return false;
  let diff = 0;
  for (let i = 0; i < a.length; i++) diff |= a.charCodeAt(i) ^ b.charCodeAt(i);
  return diff === 0;
}}

module.exports = async function handler(req, res) {{
  if (req.method === 'OPTIONS') {{ res.writeHead(204, CORS); res.end(); return; }}
  if (req.method !== 'POST')    {{ res.writeHead(405, CORS); res.end(); return; }}

  const pw = req.headers['x-settle-password'];
  const expected = process.env.SETTLE_PASSWORD;
  if (!expected) {{
    res.writeHead(503, {{ ...CORS, 'Content-Type': 'application/json' }});
    res.end(JSON.stringify({{ error: 'SETTLE_PASSWORD env var not configured' }}));
    return;
  }}
  if (!pw || !timingSafeEqualStr(String(pw), String(expected))) {{
    // delay ~250ms para mitigar timing/brute-force trivial
    await new Promise(r => setTimeout(r, 250));
    res.writeHead(401, {{ ...CORS, 'Content-Type': 'application/json' }});
    res.end(JSON.stringify({{ error: 'unauthorized' }}));
    return;
  }}

  res.writeHead(200, {{
    ...CORS,
    'Content-Type':  'application/json',
    'Cache-Control': 'no-store',
  }});
  res.end(JSON.stringify(BLOB));
}};
"""
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(src, encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help="Ruta al .xlsx de liquidación Calavo")
    ap.add_argument("--password", type=str, default=None, help="Password a usar (sino lee .settle-password o genera uno)")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"❌ No encontré: {args.xlsx}", file=sys.stderr); sys.exit(1)

    password = args.password
    if not password:
        if PWFILE.exists():
            password = PWFILE.read_text().strip()
            print(f"🔑 Reutilizando password de {PWFILE.name}")
        else:
            password = gen_password()
            PWFILE.write_text(password + "\n")
            os.chmod(PWFILE, 0o600)
            print(f"🆕 Generé password nuevo y lo guardé en {PWFILE.name} (chmod 600)")

    print(f"📄 Leyendo {args.xlsx.name}...")
    data = extract(args.xlsx)
    print(f"   · {len(data['overview_payments'])} pagos · {len(data['beltran_summary'])} líneas resumen · {len(data['embarques'])} embarques")

    print(f"🔐 Encriptando (PBKDF2 {PBKDF2_ITERS} iters → AES-256-GCM)...")
    blob = encrypt(data, password)
    write_function(blob)
    print(f"✅ Escrito {OUT.relative_to(ROOT)}")

    print()
    print("➡  Próximos pasos:")
    print(f"   1) Configura en Vercel:  vercel env add SETTLE_PASSWORD production")
    print(f"      (pega el password de {PWFILE.name})")
    print(f"   2) Commit & push: git add api/settle.js && git commit && git push")
    print(f"   3) En el dashboard, pestaña Liquidación → mete el mismo password.")


if __name__ == "__main__":
    main()
